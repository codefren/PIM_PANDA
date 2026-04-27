import openpyxl as excel
from connect_db_f import connect_db

class DB_atributos:
    def __init__(self, ruta_excel:str = "", id_inicial: int = 10) -> None:
        self.__excel = excel.load_workbook(ruta_excel)
        self.__DB = connect_db()
        self.dicccionario = {}
        self.__idatributo = id_inicial
        self.market = self.get_marketplaces() 
        

    def get_marketplaces(self) -> None:
        q = "select fldmarketplace from tbdmarketplaces"
        self.__DB.execute(q)
        res = self.__DB.fetchall()
        if res:
            return [x[0] for x in res]
        else:
            raise Exception("Can't get marketplaces")

    
    def create_all_db(self) -> dict:
        hojas = self.__excel.sheetnames
        
        # Itera a través de cada hoja
        for hoja_nombre in hojas:
            # Selecciona la hoja actual
            hoja_actual = self.__excel[hoja_nombre]

            
            # Aquí puedes realizar operaciones en la hoja actual
            # Por ejemplo, imprimir el nombre de la hoja y el contenido de la celda A1:
            print("Nombre de la hoja:", hoja_nombre)
            for i,x in enumerate(hoja_actual.iter_cols()):
                if i == 0:
                    self._familias = [fam.value for fam in x[1:] if fam.value is not None]
                else:
                    self.__idatributo += 1
                    
                    multi = "1" if len(x[0].value.split(",")) == 2 and int(x[0].value.split(",")[1].strip()) == 1 else "0"
                    nombre = x[0].value.split(",")[0]
                    
                    self.dicccionario[nombre + str(self.__idatributo)] = {"Categorias": [mod.value for mod in x[1:] if mod.value is not None],
                                                     "Familias": self._familias,
                                                     "Idatributo": str(self.__idatributo).zfill(8),
                                                     "Marketplaces": self.market,
                                                     "Multiselectable": int(multi),
                                                     "Descripcion": nombre,
                                                     "Nombre": nombre}
        
    def run(self) -> None:
        if not self.dicccionario:
            self.create_all_db()
        p1 = "select fldidatributo from tbdatributos where fldnombre = ?"
        p2 = "select fldidfamilia from tbdatributosfamilias where fldidatributo = ?"
        q1 = "insert into tbdatributos(fldidatributo,fldnombre,flddescripcion,fldmultiseleccionable) values (?,?,?,?)"
        q2 = "insert into tbdatributoscategorias(fldidatributo,fldcategoria) values(?,?)"
        q3 = "insert into tbdatributosfamilias(fldidatributo, fldidfamilia) values(?,?)"
        q4 = "insert into tbdatributosmarketplaces(fldidatributo,fldmarketplace) values (?,?)"
        
        for atributo_name in self.dicccionario:
            atributo = self.dicccionario[atributo_name]

            param1 = [atributo.get("Idatributo"), atributo.get("Nombre"), atributo.get("Descripcion"), atributo.get("Multiselectable")]

            #Primer insert
            self.__DB.execute(p1, [atributo.get("Nombre")])
            res = self.__DB.fetchone()
            if res:
                for familia in atributo.get("Familias"):
                    self.__DB.execute(p2, [res[0]])
                    res2 = self.__DB.fetchone()
                    if res2 and str(res2[0]).strip() == str(familia).strip():
                        print(f'Familia {res2[0]} ya existe')
                        continue
                    param3 = [res[0], familia]
                    self.__DB.execute(q3, param3)
                    self.__DB.commit()
                print(f"atributo {res[0]} ya existe")
                continue
            self.__DB.execute(q1, param1)
            self.__DB.commit()

            #Segundo insert
            for categoria in atributo.get("Categorias"):
                param2 = [atributo.get("Idatributo"), categoria]
                self.__DB.execute(q2, param2)
                self.__DB.commit()

            #tercer insert
            for familia in atributo.get("Familias"):
                param3 = [atributo.get("Idatributo"), familia]
                self.__DB.execute(q3, param3)
                self.__DB.commit()

            #cuarto insert
            for market in atributo.get("Marketplaces"):
                param4 = [atributo.get("Idatributo"), market]
                self.__DB.execute(q4, param4)
                self.__DB.commit()

        print("Se ha finalizado con éxito")
